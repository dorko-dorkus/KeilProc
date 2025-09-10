from __future__ import annotations

from pathlib import Path
from typing import Dict, Any
import re
import pandas as pd
import openpyxl


def _read_sheet_as_df(ws) -> pd.DataFrame:
    """Return a DataFrame from an openpyxl worksheet.

    The first row is treated as the header.  Subsequent rows form the body.
    Empty worksheets yield an empty DataFrame.
    """
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    data = rows[1:]
    return pd.DataFrame(data, columns=header)


def _scale_raw_with_alpha_beta(df: pd.DataFrame):
    """Scale a raw piccolo series using alpha/beta columns if present.

    This helper looks for columns named ``Piccolo_raw`` together with
    ``alpha`` and ``beta`` values.  If found, the scaled series
    ``alpha * raw + beta`` is returned; otherwise ``None`` is returned.
    """
    if {"Piccolo_raw", "alpha", "beta"} <= set(df.columns):
        try:
            raw = pd.to_numeric(df["Piccolo_raw"], errors="coerce")
            alpha = float(pd.to_numeric(df["alpha"], errors="coerce").iloc[0])
            beta = float(pd.to_numeric(df["beta"], errors="coerce").iloc[0])
            return alpha * raw + beta
        except Exception:
            return None
    return None


def extract_piccolo_overlay_from_workbook(xlsx_path: Path, out_csv: Path) -> Dict[str, Any]:
    """Extract a piccolo differential pressure series from a legacy workbook.

    Multiple heuristics are attempted in order.  The first successful series is
    written to ``out_csv`` with a ``DP_mbar`` column and meta information about
    the extraction is returned.  If no series can be located, ``status`` is set
    to ``"no_dp_found"``.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    # ---- (0) Read Piccolo range from 'Data' sheet if present ----
    range_mbar = None
    if "Data" in wb.sheetnames:
        rows = list(wb["Data"].values)[:200]
        for r in rows:
            vals = [str(v).strip() if v is not None else "" for v in r]
            if any("piccolo tx range setting" in v.lower() for v in vals):
                for v in r:
                    if isinstance(v, (int, float)):
                        range_mbar = float(v)
                        break
                if range_mbar is not None:
                    break

    # ---- (1) Try explicit DP / Piccolo_eng / alpha-beta paths (unchanged) ----
    for ws in wb.worksheets:
        df = _read_sheet_as_df(ws)
        if df.empty:
            continue
        for col in df.columns:
            nm = str(col).strip().lower()
            if re.fullmatch(r"dp(_mbar)?|(dp \(mbar\))", nm):
                out = pd.DataFrame({"DP_mbar": pd.to_numeric(df[col], errors="coerce")})
                out_csv.parent.mkdir(parents=True, exist_ok=True)
                out.to_csv(out_csv, index=False)
                return {"status": "ok", "source": f"explicit:{col}", "rows": int(out.shape[0])}
            if nm.startswith("piccolo_eng"):
                out = pd.DataFrame({"DP_mbar": pd.to_numeric(df[col], errors="coerce")})
                out_csv.parent.mkdir(parents=True, exist_ok=True)
                out.to_csv(out_csv, index=False)
                return {"status": "ok", "source": f"Piccolo_eng:{col}", "rows": int(out.shape[0])}
        dp_series = _scale_raw_with_alpha_beta(df)
        if dp_series is not None:
            out = pd.DataFrame({"DP_mbar": pd.to_numeric(dp_series, errors="coerce")})
            out_csv.parent.mkdir(parents=True, exist_ok=True)
            out.to_csv(out_csv, index=False)
            return {"status": "ok", "source": "Piccolo_raw_scaled_alpha_beta", "rows": int(out.shape[0])}

    # ---- (2) Piccolo Tx Current (mA) + Range (mbar) on P# sheets ----
    if range_mbar is not None:
        series = []
        for name in wb.sheetnames:
            if not re.fullmatch(r"P\d+", name):
                continue
            ws = wb[name]
            rows = list(ws.values)
            header_idx = None
            header = None
            for i, r in enumerate(rows[:120]):
                vals = [str(v).strip() if v is not None else "" for v in r]
                if "Time" in vals and "Piccolo Tx Current" in vals:
                    header_idx = i
                    header = vals
                    break
            if header_idx is None:
                continue
            df = pd.DataFrame(rows[header_idx + 1 :], columns=header)
            cur = pd.to_numeric(df.get("Piccolo Tx Current"), errors="coerce")
            good = cur.notna()
            sub = df.loc[good, ["Time", "Piccolo Tx Current"]].copy()
            sub["DP_mbar"] = (pd.to_numeric(sub["Piccolo Tx Current"], errors="coerce") - 4.0) / 16.0 * float(range_mbar)
            series.append(sub[["Time", "DP_mbar"]])
        if series:
            out = pd.concat(series, ignore_index=True)
            out = out[out["Time"].astype(str).str.lower() != "averages"]
            out_csv.parent.mkdir(parents=True, exist_ok=True)
            out.to_csv(out_csv, index=False)
            return {
                "status": "ok",
                "source": "Piccolo_Tx_Current_4-20mA_with_Range",
                "rows": int(out.shape[0]),
            }

    return {"status": "no_dp_found", "source": None, "rows": 0}


# ---- Helper: Piccolo range and average Tx current from Data sheet ----
def extract_piccolo_range_and_avg_current(xlsx_path: Path) -> Dict[str, Any]:
    """Scrape piccolo range and average transmit current from the Data sheet.

    The function searches the ``Data`` worksheet for rows containing the
    labels ``Piccolo Tx Range`` and ``Average Piccolo Tx Current``.  Units are
    normalised to mbar for the range and mA for the current.  If the sheet is
    absent or the values cannot be located, ``None`` is returned for the
    corresponding fields.

    Returns
    -------
    dict
        ``{"status": "...", "range_mbar": float|None, "avg_current_mA": float|None}``
    """

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "Data" not in wb.sheetnames:
            return {"status": "absent", "range_mbar": None, "avg_current_mA": None}

        ws = wb["Data"]
        rng = None
        avg_i = None

        # Scan the first ~200 rows for the two labels
        rows = list(ws.values)[:200]
        for r in rows:
            vals = [str(v).strip() if v is not None else "" for v in r]
            line = " ".join(vals).lower()

            if "piccolo tx range" in line:
                for v in r:
                    if isinstance(v, (int, float)):
                        rng = float(v)
                        break
                unit = next(
                    (s for s in vals if s.lower() in ("mbar", "kpa", "pa", "mb")),
                    "mbar",
                ).lower()
                if unit == "kpa":
                    rng = rng * 10.0  # 1 kPa = 10 mbar
                if unit == "pa":
                    rng = rng / 100.0  # 100 Pa = 1 mbar

            if "average piccolo tx current" in line:
                for v in r:
                    if isinstance(v, (int, float)):
                        avg_i = float(v)
                        break

        return {"status": "ok", "range_mbar": rng, "avg_current_mA": avg_i}
    except Exception as e:  # pragma: no cover - protective
        return {
            "status": "error",
            "error": str(e),
            "range_mbar": None,
            "avg_current_mA": None,
        }


# --- NEW: extract barometric pressure (Data!H15:I19) with units ---
def extract_baro_from_workbook(xlsx_path: Path) -> Dict[str, Any]:
    """
    Read barometric pressure from the legacy workbook's Data sheet.
    Expected region: H15:I19 where column H is a unit label and column I is the value.
    Typical rows:
        H15='kPa'   I15=101.6
        H16='C'     I16=<dry-bulb>
        H17='C'     I17=<wet-bulb or other>
        H18='kg/m3' I18=<density>
        H21='m'     I21=<elevation>
    Returns:
        {"status":"ok","baro_pa":float,"unit_raw":str,"cell":"Data!I15"} or {"status":"absent"}.
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "Data" not in wb.sheetnames:
            return {"status": "absent"}
        ws = wb["Data"]
        # Scan H15:I19 (rows 15..19, cols H=8, I=9)
        for r in range(15, 20):
            unit = ws.cell(row=r, column=8).value  # H
            val = ws.cell(row=r, column=9).value  # I
            if val is None or unit is None:
                continue
            unit_s = str(unit).strip().lower()
            # Map known units -> Pa
            if isinstance(val, (int, float)):
                if unit_s in ("kpa",):
                    pa = float(val) * 1000.0
                elif unit_s in ("pa",):
                    pa = float(val)
                elif unit_s in ("mbar", "mb"):
                    pa = float(val) * 100.0
                else:
                    # skip rows that are clearly not pressure (e.g., 'c', 'kg/m3', 'm')
                    continue
                if pa > 0:
                    return {
                        "status": "ok",
                        "baro_pa": pa,
                        "unit_raw": unit_s,
                        "cell": f"Data!I{r}",
                    }
        return {"status": "absent"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


# --- NEW: dry-bulb temperature (Data!H15:I19) in °C → K ---
def extract_temperature_from_workbook(xlsx_path: Path) -> Dict[str, Any]:
    """Read dry-bulb temperature from a legacy workbook's Data sheet.

    The function scans ``Data!H15:I19`` for a row where column ``H`` contains
    a Celsius unit (``C``, ``°C``, ``degC``, ``deg C``) and column ``I`` holds a
    numeric value.  The value is converted to Kelvin.

    Returns
    -------
    dict
        ``{"status": "ok", "T_K": float, "cell": "Data!I.."}`` when a value
        is found, ``{"status": "absent"}`` when the region lacks such a row,
        or ``{"status": "error", "error": str}`` on exceptions.
    """

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "Data" not in wb.sheetnames:
            return {"status": "absent"}
        ws = wb["Data"]
        for r in range(15, 20):  # H15..I19
            unit = ws.cell(row=r, column=8).value  # H
            val = ws.cell(row=r, column=9).value  # I
            if unit is None or val is None:
                continue
            unit_s = str(unit).strip().lower()
            if unit_s in ("c", "°c", "degc", "deg c"):
                try:
                    T_K = float(val) + 273.15
                except Exception:
                    continue
                return {"status": "ok", "T_K": T_K, "cell": f"Data!I{r}"}
        return {"status": "absent"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


def extract_process_temperature_from_workbook(xlsx_path: Path) -> dict:
    """Extract PRIMARY AIR / PROCESS temperature in °C (convert to K).

    Heuristics
    ---------
    * Scan the first ~200 rows of the ``Data`` sheet for a unit cell
      indicating Celsius (``C``, ``°C``, ``degC``, ``deg C``) with a numeric
      value nearby.
    * Prefer rows whose label mentions any of
      ``['primary air', 'pa', 'process', 'mill', 'duct', 'gas']``.
    * If multiple candidates exist, prefer temperatures ``> 80°C`` to avoid
      picking ambient values, otherwise choose the hottest.

    Returns
    -------
    dict
        ``{"status":"ok","T_K": float, "cell": "Data!I.."}`` when a value
        is found, ``{"status":"absent"}`` when no suitable row is located,
        or ``{"status":"error", "error": str}`` on failure.
    """

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "Data" not in wb.sheetnames:
            return {"status": "absent"}
        ws = wb["Data"]
        rows = list(
            ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=12, values_only=True)
        )
        cand = []
        for ridx, row in enumerate(rows, start=1):
            for cidx, val in enumerate(row):
                unit = str(val).strip().lower() if val is not None else ""
                if unit in ("c", "°c", "degc", "deg c"):
                    v = None
                    vcol = None
                    if cidx + 1 < len(row) and isinstance(row[cidx + 1], (int, float)):
                        v = float(row[cidx + 1])
                        vcol = cidx + 1
                    else:
                        for j, x in enumerate(row):
                            if isinstance(x, (int, float)):
                                v = float(x)
                                vcol = j
                                break
                    if v is None:
                        continue
                    label = " ".join([str(x) for x in row if isinstance(x, str)]).lower()
                    score = 0
                    for kw in ("primary air", " pa ", "process", "mill", "duct", "gas"):
                        if kw in (" " + label + " "):
                            score += 10
                    if v >= 80.0:
                        score += 5
                    cand.append(
                        {
                            "score": score,
                            "T_C": v,
                            "cell": f"Data!{chr(65 + vcol)}{ridx}",
                        }
                    )
        if not cand:
            return {"status": "absent"}
        cand.sort(key=lambda x: (x["score"], x["T_C"]))
        best = cand[-1]
        return {
            "status": "ok",
            "T_K": best["T_C"] + 273.15,
            "cell": best["cell"],
        }
    except Exception as e:  # pragma: no cover - protective
        return {"status": "error", "error": str(e)}


__all__ = [
    "extract_piccolo_overlay_from_workbook",
    "extract_baro_from_workbook",
    "extract_temperature_from_workbook",
    "extract_process_temperature_from_workbook",
    "extract_piccolo_range_and_avg_current",
]
