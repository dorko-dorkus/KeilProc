from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Dict, Any, Tuple, List
import json, math, re
import numpy as np
import pandas as pd
import openpyxl

# ---------- Season presets ----------
# Source of truth for 820 linearization (flow = m*DP + c).
# You can override these via site.defaults["calib_820_summer"] / ["calib_820_winter"]
# or by providing season-specific calibration workbooks (see _calib_for_season()).
SEASON_PRESETS = {
    "summer": {"m": 8.40, "c": 31.50},   # t/h per mbar, t/h  (example values)
    "winter": {"m": 8.40, "c": 31.50},   # TODO: update when winter fit available
}
# UIC constant (Flow_UIC = K * sqrt(DP_mbar))
K_UIC_DEFAULT = 33.50   # t/h per sqrt(mbar)  (example value from your workbook)

@dataclass
class SeasonCalib:
    season: str
    K_uic: float
    m_820: float
    c_820: float
    source: str

# ---------- Helpers ----------
def _fit_from_workbook(xlsx_path: Path) -> Tuple[float, float, float]:
    """Fit (K, m, c) from Sheet1 table: col2=DP_mbar, col3=UIC_flow, col4=820_flow."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    dps: List[float] = []; uic: List[float] = []; lin: List[float] = []
    for r in range(19, 500):
        dp = ws.cell(row=r, column=2).value
        f3 = ws.cell(row=r, column=3).value
        f4 = ws.cell(row=r, column=4).value
        if isinstance(dp, (int, float)) and isinstance(f3, (int, float)) and isinstance(f4, (int, float)) and dp >= 0:
            dps.append(float(dp)); uic.append(float(f3)); lin.append(float(f4))
    if not dps:
        raise ValueError("No DP rows found in calibration workbook (Sheet1 col2..4).")
    dps = np.array(dps); uic = np.array(uic); lin = np.array(lin)
    K = np.median(uic[dps > 0] / np.sqrt(dps[dps > 0]))
    A = np.vstack([dps, np.ones_like(dps)]).T
    m, c = np.linalg.lstsq(A, lin, rcond=None)[0]
    return float(K), float(m), float(c)

def _calib_for_season(season: str,
                      site_defaults: Optional[dict] = None,
                      season_workbooks: Optional[dict] = None) -> SeasonCalib:
    s = (season or "summer").lower()
    # 1) workbook override for this season
    if season_workbooks and season_workbooks.get(s):
        K, m, c = _fit_from_workbook(Path(season_workbooks[s]))
        return SeasonCalib(s, K, m, c, source=f"workbook:{season_workbooks[s]}")
    # 2) site.defaults override (explicit m/c/K)
    if site_defaults:
        mck = site_defaults.get(f"calib_820_{s}")  # e.g., {"m": 8.4, "c": 31.5, "K_uic": 33.5}
        if isinstance(mck, dict) and "m" in mck and "c" in mck:
            return SeasonCalib(s, float(mck.get("K_uic", K_UIC_DEFAULT)), float(mck["m"]), float(mck["c"]), "site.defaults")
    # 3) built-in presets
    preset = SEASON_PRESETS.get(s, SEASON_PRESETS["summer"])
    return SeasonCalib(s, K_UIC_DEFAULT, float(preset["m"]), float(preset["c"]), "preset")

def _autodetect_dp_col(df: pd.DataFrame) -> Tuple[str, str]:
    """Return (dp_col, inferred_unit). Infer unit by magnitude if needed."""
    name_candidates = ["dp_mbar", "dp (mbar)", "dp", "i/p", "diff", "differential"]
    for nm in df.columns:
        if any(re.fullmatch(pat.replace(" ", r"\s*"), nm.strip().lower()) for pat in name_candidates):
            series = df[nm].astype(float, errors="ignore")
            return nm, _infer_unit_from_series(series)
    # fallback: first numeric column
    for nm in df.columns:
        try:
            series = pd.to_numeric(df[nm], errors="coerce")
            if series.notna().sum() > 0:
                return nm, _infer_unit_from_series(series)
        except Exception:
            continue
    raise ValueError("No numeric DP-like column found in logger CSV.")

def _infer_unit_from_series(s: pd.Series) -> str:
    s = pd.to_numeric(s, errors="coerce").dropna()
    if s.empty:
        return "mbar"
    mx = float(s.quantile(0.99))
    # crude heuristics: PA typically thousands, mbar typically < 100
    if mx >= 500: return "Pa"
    return "mbar"

def _to_mbar(values: pd.Series, unit_hint: str) -> pd.Series:
    u = (unit_hint or "mbar").lower()
    if u in ("mbar","mb","millibar","milli-bar"):
        return pd.to_numeric(values, errors="coerce")
    if u in ("pa","pascal","pascals"):
        return pd.to_numeric(values, errors="coerce") / 100.0
    if u in ("kpa",):
        return pd.to_numeric(values, errors="coerce") * 10.0
    return pd.to_numeric(values, errors="coerce")

# ---------- Builders ----------
def build_reference_table(cal: SeasonCalib, *, dp_max_mbar: float = 10.0, dp_step_mbar: float = 0.1) -> pd.DataFrame:
    grid = np.arange(0.0, dp_max_mbar + 1e-9, dp_step_mbar)
    flow_uic = cal.K_uic * np.sqrt(grid)
    flow_820 = cal.m_820 * grid + cal.c_820
    return pd.DataFrame({
        "ref_DP_mbar": grid,
        "ref_Flow_UIC_tph": flow_uic,
        "ref_Flow_820_tph": flow_820,
        "ref_Flow_err_820_minus_UIC_tph": flow_820 - flow_uic,
    })

def build_data_overlay(cal: SeasonCalib, df_logger: pd.DataFrame, dp_col: Optional[str] = None, dp_unit_hint: Optional[str] = None) -> pd.DataFrame:
    if dp_col is None:
        dp_col, inferred = _autodetect_dp_col(df_logger)
        dp_unit_hint = dp_unit_hint or inferred
    dp_mbar = _to_mbar(df_logger[dp_col], dp_unit_hint)
    flow_uic = cal.K_uic * np.sqrt(np.clip(dp_mbar.values, 0.0, None))
    flow_820 = cal.m_820 * dp_mbar.values + cal.c_820
    out = pd.DataFrame({
        "data_DP_mbar": dp_mbar.values,
        "data_Flow_UIC_tph": flow_uic,
        "data_Flow_820_tph": flow_820,
        "data_Flow_err_820_minus_UIC_tph": flow_820 - flow_uic,
    })
    return out

def write_lookup_outputs(
    outdir: Path,
    *,
    season: str,
    site_defaults: Optional[dict] = None,
    season_workbooks: Optional[dict] = None,   # {"summer": "/path.xlsx", "winter": "/path.xlsx"}
    logger_csv: Optional[Path] = None,
    dp_col: Optional[str] = None,
    dp_unit_hint: Optional[str] = None,
    dp_max_mbar: Optional[float] = None,
    dp_step_mbar: float = 0.1,
) -> Dict[str, Any]:
    """Always write the constant reference table; add data overlay if logger present."""
    outdir = Path(outdir); outdir.mkdir(parents=True, exist_ok=True)
    cal = _calib_for_season(season, site_defaults=site_defaults, season_workbooks=season_workbooks)
    # reference side (constant)
    ref = build_reference_table(cal, dp_max_mbar=float(dp_max_mbar or 10.0), dp_step_mbar=dp_step_mbar)
    ref_csv = outdir / "transmitter_lookup_reference.csv"
    ref.to_csv(ref_csv, index=False)
    # data side (optional)
    overlay_csv = None; combined_csv = None
    overlay = None
    if logger_csv and Path(logger_csv).exists():
        df_log = pd.read_csv(logger_csv)
        overlay = build_data_overlay(cal, df_log, dp_col=dp_col, dp_unit_hint=dp_unit_hint)
        overlay_csv = outdir / "transmitter_lookup_data.csv"
        overlay.to_csv(overlay_csv, index=False)
        # combined view (constant left, data right)
        combined = pd.concat([ref, overlay], axis=1)
        combined_csv = outdir / "transmitter_lookup_combined.csv"
        combined.to_csv(combined_csv, index=False)
    meta = {
        "season": cal.season,
        "calibration": {"K_uic": cal.K_uic, "m_820": cal.m_820, "c_820": cal.c_820, "source": cal.source},
        "reference_csv": str(ref_csv),
        "overlay_csv": (str(overlay_csv) if overlay_csv else None),
        "combined_csv": (str(combined_csv) if combined_csv else None),
    }
    (outdir / "transmitter_lookup_meta.json").write_text(json.dumps(meta, indent=2))
    return meta
