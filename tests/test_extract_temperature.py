from pathlib import Path

from openpyxl import Workbook

from kielproc.tools.legacy_overlay import (
    extract_temperature_from_workbook,
    extract_process_temperature_from_workbook,
)


def test_extract_temperature_from_workbook(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["H15"] = "C"
    ws["I15"] = 20.0
    f = tmp_path / "wb.xlsx"
    wb.save(f)

    res = extract_temperature_from_workbook(Path(f))
    assert res["status"] == "ok"
    assert abs(res["T_K"] - (20.0 + 273.15)) < 1e-6
    assert res["cell"] == "Data!I15"
    assert res["value"] == 20.0
    assert res["unit"] == "C"


def test_extract_process_temperature_from_workbook(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Ambient", "C", 25.0])
    ws.append(["Primary Air Temp", "C", 150.0])
    f = tmp_path / "wb.xlsx"
    wb.save(f)

    res = extract_process_temperature_from_workbook(Path(f))
    assert res["status"] == "ok"
    assert abs(res["T_K"] - (150.0 + 273.15)) < 1e-6
    assert res["cell"] == "Data!C2"
    assert res["value"] == 150.0
    assert res["unit"] == "C"

