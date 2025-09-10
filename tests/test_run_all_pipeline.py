from pathlib import Path
import json
import pandas as pd

from kielproc.run_easy import RunConfig, SitePreset, run_all


def test_run_all_produces_outputs(tmp_path):
    in_dir = tmp_path / "in"
    out_dir = tmp_path / "out"
    in_dir.mkdir()
    out_dir.mkdir()

    # Minimal port CSV (P1)
    pd.DataFrame(
        {
            "VP": [10.0, 12.0],
            "Temperature": [25.0, 26.0],
            "Static": [101325.0, 101300.0],
        }
    ).to_csv(in_dir / "run__P1.csv", index=False)

    # Logger CSV for transmitter setpoints and flow lookup
    sp_csv = tmp_path / "setpoints.csv"
    pd.DataFrame({"DP_mbar": [10, 20], "T_C": [25, 30]}).to_csv(sp_csv, index=False)

    cfg = RunConfig(
        input_dir=str(in_dir),
        output_dir=str(out_dir),
        enable_site=True,
        site=SitePreset(
            name="SiteA",
            geometry={"duct_width_m": 1.0, "duct_height_m": 1.0, "throat_area_m2": 0.25},
        ),
        setpoints_csv=str(sp_csv),
    )

    summary = run_all(cfg)

    # Basic sanity checks
    assert Path(summary["per_port_csv"]).exists()
    assert Path(summary["duct_result_json"]).exists()
    assert summary["input_mode"] == "csv_folder"
    assert Path(summary["prepared_input_dir"]) == in_dir
    tp = out_dir / "_integrated" / "transmitter_setpoints.csv"
    assert tp.exists()
    assert summary["setpoints"]["rows"] == 2
    # Flow lookup artifacts
    ref = out_dir / "_integrated" / "transmitter_lookup_reference.csv"
    combined = out_dir / "_integrated" / "transmitter_lookup_combined.csv"
    assert ref.exists()
    assert combined.exists()
    overlay = Path(summary["flow_lookup"]["overlay_csv"])
    assert overlay.exists()
    assert Path(summary["flow_lookup"]["combined_csv"]).exists()
    assert summary["baro_pa"] == 101_325.0
    assert summary["site_name"] == "SiteA"
    # Venturi result files present and sane
    vr = Path(summary["venturi_result_json"])
    assert vr.exists()
    data = json.loads(vr.read_text())
    assert len(data["curve"]) == 10
    assert data["dp_vent_Pa_at_Qs"] > 0


def test_run_all_accepts_workbook(tmp_path):
    wb_path = tmp_path / "input.xlsx"
    out_dir = tmp_path / "out"
    out_dir.mkdir()

    # Build workbook with blank unit row
    cols = ["Time", "Static Pressure", "Velocity Pressure", "Temperature", "Piccolo", "Piccolo Tx Current"]
    data = [[0, 101325, 10, 25, 0.1, 8.0], [1, 101300, 12, 26, 0.2, 12.0]]
    with pd.ExcelWriter(wb_path, engine="openpyxl") as writer:
        pd.DataFrame(data, columns=cols).to_excel(writer, sheet_name="P1", index=False)
        pd.DataFrame([["Piccolo Tx Range Setting", 6.7]]).to_excel(writer, sheet_name="Data", index=False, header=False)
    from openpyxl import load_workbook

    wb = load_workbook(wb_path)
    wb["P1"].insert_rows(2)
    ws = wb["Data"]
    ws["H15"] = "kPa"
    ws["I15"] = 101.6
    wb.save(wb_path)

    cfg = RunConfig(
        input_dir=str(wb_path),
        output_dir=str(out_dir),
        enable_site=True,
        site=SitePreset(
            name="SiteA",
            geometry={"duct_width_m": 1.0, "duct_height_m": 1.0, "throat_area_m2": 0.25},
        ),
    )

    summary = run_all(cfg)

    assert summary["input_mode"] == "legacy_workbook"
    prepared = Path(summary["prepared_input_dir"])
    assert prepared.exists()
    # per-port CSV generated from workbook and consumed
    assert Path(summary["per_port_csv"]).exists()
    # Flow lookup reference uses the extracted overlay
    meta = summary["flow_lookup"]
    assert Path(meta["reference_csv"]).exists()
    assert Path(meta["overlay_csv"]).exists()
    picc = summary["piccolo_overlay"]
    assert picc["status"] == "ok"
    csv = Path(picc["csv"])
    assert csv.exists()
    dfp = pd.read_csv(csv)
    assert abs(dfp["DP_mbar"].iloc[0] - ((8.0 - 4.0) / 16.0 * 6.7)) < 1e-6
    assert abs(dfp["DP_mbar"].iloc[1] - ((12.0 - 4.0) / 16.0 * 6.7)) < 1e-6
    # Barometric pressure extracted from workbook Data sheet
    assert summary["baro_pa"] == 101_600.0
    assert summary["baro"]["status"] == "ok"
